"""
CCTV Daily Check Bot v9
========================
Changes from v8:
  - Improved clarity check:
    1. Center-region sharpness (not whole image)
    2. Timestamp visibility check (white pixels in top-left)
    3. Lens blocked/covered detection (very low color variance)
    4. Brightness range check (overexposed / underlit)
    5. Night vision detection (grayscale IR image)
  - Date/Time: DVR system time API (100% accurate, no OCR)
  - Excel: Summary sheet + All Cameras single sheet with auto-filter

Requirements:
    pip install requests openpyxl urllib3 pillow opencv-python
"""

import requests, json, os, sys, re
import xml.etree.ElementTree as ET
from datetime import datetime, date
from requests.auth import HTTPDigestAuth, HTTPBasicAuth
import urllib3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import cv2, numpy as np
    CV2_OK = True
except ImportError:
    CV2_OK = False

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(SCRIPT_DIR, 'nvr_config.json')
TODAY_STR   = date.today().strftime('%Y-%m-%d')
RPT_DATE    = date.today().strftime('%d-%m-%Y')
SNAP_DIR    = os.path.join(os.path.expanduser('~'), 'Desktop', 'CCTV_Snapshots', TODAY_STR)
EXCEL_PATH  = os.path.join(os.path.expanduser('~'), 'Desktop', f'CCTV_Report_{RPT_DATE}.xlsx')
TIMEOUT     = 12

C = {
    'header'  : '1F4E79', 'header_fg': 'FFFFFF',
    'ok_bg'   : 'E2EFDA', 'ok_fg'   : '375623',
    'ok_alt'  : 'D6F0C8',
    'warn_bg' : 'FFECC6', 'warn_fg' : '7F4F00',
    'err_bg'  : 'FFDCE1', 'err_fg'  : '9C0006',
    'blue_hdr': '2E75B6', 'dvr_bg'  : 'D6E4F0',
    'grey_bg' : 'F5F5F5',
}

# ── XML helper ─────────────────────────────────────────────────
def xfind(el, tag):
    v = el.find(tag)
    if v is not None: return v.text.strip() if v.text else ''
    for ns in ['{http://www.hikvision.com/ver20/XMLSchema}',
               '{http://www.isapi.org/ver20/XMLSchema}']:
        v = el.find(f'{ns}{tag}')
        if v is not None: return v.text.strip() if v.text else ''
    for child in el.iter():
        if (child.tag.split('}')[-1] if '}' in child.tag else child.tag) == tag:
            return child.text.strip() if child.text else ''
    return None

def xfindall(el, tag):
    res = []
    for child in el.iter():
        if (child.tag.split('}')[-1] if '}' in child.tag else child.tag) == tag:
            res.append(child)
    return res

def parse_hik_time(s):
    if not s: return None
    s = re.sub(r'[+-]\d{2}:\d{2}$', '', s.strip()).replace('Z','').strip()[:19]
    try: return datetime.strptime(s, '%Y-%m-%dT%H:%M:%S')
    except: return None


# ══════════════════════════════════════════════════════════════
# CLARITY CHECK — 5 checks matching what you do manually
# ══════════════════════════════════════════════════════════════

def check_clarity(img_bytes):
    """
    5-check clarity analysis matching manual inspection:

    Check 1 — Snapshot exists and is valid
    Check 2 — Brightness (black screen / very dark / overexposed)
    Check 3 — Timestamp visible (white pixels in top-left corner)
    Check 4 — Lens blocked/covered (extremely low color variance)
    Check 5 — Sharpness on CENTER region only (ignores soft edges)

    Returns (status, detail_string)

    Status values:
      CLEAR          — all checks passed
      NO SNAPSHOT    — failed to fetch image
      NO VIDEO       — black screen (camera offline)
      VERY DARK      — too dark to see (night, light off)
      NIGHT VISION   — IR mode active (grayscale), image otherwise ok
      OVEREXPOSED    — too bright, washed out
      LENS BLOCKED   — lens covered or pointing at wall
      BLURRY         — image not sharp (dirty lens / out of focus)
      NO TIMESTAMP   — camera feed ok but date/time overlay not visible
    """
    # ── Check 1: Snapshot exists ──────────────────────────────
    if not img_bytes or len(img_bytes) < 3000:
        return 'NO SNAPSHOT', 'Image too small or missing'

    if not CV2_OK:
        # Fallback without OpenCV — just size check
        return ('CLEAR' if len(img_bytes) > 10000 else 'POSSIBLY DARK',
                f'{len(img_bytes)} bytes')

    try:
        arr = cv2.imdecode(np.frombuffer(img_bytes, np.uint8), cv2.IMREAD_COLOR)
        if arr is None:
            return 'NO SNAPSHOT', 'Cannot decode image'

        h, w   = arr.shape[:2]
        gray   = cv2.cvtColor(arr, cv2.COLOR_BGR2GRAY)

        # ── Check 2: Brightness ───────────────────────────────
        mean_bri = float(np.mean(gray))
        max_bri  = float(np.max(gray))
        min_bri  = float(np.min(gray))

        if mean_bri < 12:
            return 'NO VIDEO', f'Brightness:{mean_bri:.0f} (black screen)'
        if mean_bri < 38:
            return 'VERY DARK', f'Brightness:{mean_bri:.0f}'
        if mean_bri > 240 and min_bri > 200:
            return 'OVEREXPOSED', f'Brightness:{mean_bri:.0f} (washed out)'

        # ── Check 3: Night vision detection ───────────────────
        # IR cameras: all 3 BGR channels nearly equal = grayscale image
        b_mean = float(np.mean(arr[:,:,0]))
        g_mean = float(np.mean(arr[:,:,1]))
        r_mean = float(np.mean(arr[:,:,2]))
        channel_diff = max(abs(b_mean-g_mean), abs(g_mean-r_mean), abs(b_mean-r_mean))
        is_night_vision = (channel_diff < 6.0)  # all channels nearly equal

        # ── Check 4: Lens blocked / covered ──────────────────
        # If image is uniform color (very low std dev) = blocked lens or wall
        std_dev = float(np.std(gray))
        if std_dev < 8.0 and mean_bri > 12:
            return 'LENS BLOCKED', f'StdDev:{std_dev:.1f} (uniform image)'

        # ── Check 5: Sharpness on CENTER region only ──────────
        # Take center 50% of image — avoids soft edges of wide-angle cameras
        cy1 = h // 4
        cy2 = 3 * h // 4
        cx1 = w // 4
        cx2 = 3 * w // 4
        center = gray[cy1:cy2, cx1:cx2]

        lap_var    = float(cv2.Laplacian(center, cv2.CV_64F).var())
        edges      = cv2.Canny(center, 50, 150)
        edge_pct   = float(np.count_nonzero(edges)) / (center.size) * 100

        # ── Check 6: Timestamp visible in top-left ────────────
        # Timestamp area: top 10% height, left 40% width
        ts_region  = gray[:int(h * 0.10), :int(w * 0.40)]
        # White pixels (brightness > 200) indicate timestamp text
        white_px   = np.sum(ts_region > 200)
        ts_total   = ts_region.size
        white_pct  = white_px / ts_total * 100
        ts_visible = white_pct > 0.3   # at least 0.3% white pixels

        detail = (f'Bri:{mean_bri:.0f} Sharp:{lap_var:.0f} '
                  f'Edge:{edge_pct:.1f}% TS:{"OK" if ts_visible else "NO"}'
                  f'{" IR" if is_night_vision else ""}')

        # Blurry check — needs BOTH low sharpness AND low edges
        # This avoids flagging plain-room cameras as blurry
        if lap_var < 60 and edge_pct < 1.0:
            return 'BLURRY', detail
        if lap_var < 35:
            return 'BLURRY', detail

        # Night vision — note it but don't flag as error
        if is_night_vision:
            if not ts_visible:
                return 'NIGHT VISION - NO TIMESTAMP', detail
            return 'NIGHT VISION', detail

        # Timestamp not visible
        if not ts_visible:
            return 'NO TIMESTAMP', detail

        return 'CLEAR', detail

    except Exception as e:
        return 'CHECK ERROR', str(e)


# ══════════════════════════════════════════════════════════════
class DVRChecker:
    def __init__(self, cfg):
        self.name          = cfg['name']
        self.ip            = cfg['ip_address']
        self.port          = cfg.get('port', 80)
        self.user          = cfg['username']
        self.pwd           = cfg['password']
        self.base          = f"http://{self.ip}:{self.port}"
        self.auth          = HTTPDigestAuth(self.user, self.pwd)
        self.skip_channels = set(cfg.get('skip_channels', []))

    def get(self, path, stream=False):
        url = f"{self.base}{path}"
        for auth in [HTTPDigestAuth(self.user, self.pwd),
                     HTTPBasicAuth(self.user, self.pwd)]:
            try:
                r = requests.get(url, auth=auth, verify=False,
                                 timeout=TIMEOUT, stream=stream)
                if r.status_code == 401: continue
                self.auth = auth
                return r, None
            except requests.exceptions.ConnectTimeout: return None, 'TIMEOUT'
            except requests.exceptions.ConnectionError: return None, 'UNREACHABLE'
            except Exception as e: return None, str(e)
        return None, 'AUTH_FAILED'

    def post(self, path, body):
        try:
            r = requests.post(f"{self.base}{path}", data=body, auth=self.auth,
                              headers={'Content-Type': 'application/xml'},
                              verify=False, timeout=TIMEOUT)
            return r, None
        except Exception as e: return None, str(e)

    # ── System time ────────────────────────────────────────────
    def get_system_time(self):
        r, err = self.get('/ISAPI/System/time')
        if err: return None, err
        if r.status_code != 200: return None, f'HTTP {r.status_code}'
        try: root = ET.fromstring(r.text)
        except ET.ParseError as e: return None, f'XML:{e}'
        t = xfind(root, 'localTime')
        if not t: return None, 'localTime not found'
        dt = parse_hik_time(t)
        if not dt: return None, f'Cannot parse: {t}'
        return dt, 'OK'

    # ── Camera names ───────────────────────────────────────────
    def get_camera_names(self):
        names = {}
        r, _ = self.get('/ISAPI/System/Video/inputs/channels')
        if r and r.status_code == 200:
            try:
                root = ET.fromstring(r.text)
                for ch in xfindall(root, 'VideoInputChannel'):
                    cid  = xfind(ch, 'id')
                    name = xfind(ch, 'n') or xfind(ch, 'name')
                    if cid and name: names[int(cid)] = name
            except: pass
        if not names:
            r, _ = self.get('/ISAPI/ContentMgmt/InputProxy/channels')
            if r and r.status_code == 200:
                try:
                    root = ET.fromstring(r.text)
                    for ch in xfindall(root, 'InputProxyChannel'):
                        cid  = xfind(ch, 'id')
                        name = xfind(ch, 'name')
                        if cid and name: names[int(cid)] = name
                except: pass
        return names

    # ── Snapshot ───────────────────────────────────────────────
    def get_snapshot(self, ch):
        tid = ch * 100 + 1
        for path in [f'/ISAPI/Streaming/channels/{tid}/picture',
                     f'/ISAPI/Streaming/channels/{ch}/picture']:
            r, err = self.get(path, stream=True)
            if err: continue
            if r and r.status_code == 200 and len(r.content) > 2000:
                return r.content, None
        return None, 'SNAPSHOT_FAILED'

    # ── Recording check ────────────────────────────────────────
    def check_recording(self, ch):
        tid   = ch * 100 + 1
        today = date.today().strftime('%Y-%m-%d')
        for tz in ['+05:30', 'Z']:
            body = (f'<CMSearchDescription><searchID>{ch}</searchID>'
                    f'<trackList><trackID>{tid}</trackID></trackList>'
                    f'<timeSpanList><timeSpan>'
                    f'<startTime>{today}T00:00:01{tz}</startTime>'
                    f'<endTime>{today}T23:59:59{tz}</endTime>'
                    f'</timeSpan></timeSpanList>'
                    f'<maxResults>1</maxResults><searchResultPostion>0</searchResultPostion>'
                    f'<metadataList><metadataDescriptor>'
                    f'//recordType.meta.hikvision.com/dataType'
                    f'</metadataDescriptor></metadataList></CMSearchDescription>')
            r, _ = self.post('/ISAPI/ContentMgmt/search', body)
            if r and r.status_code == 200 and 'numOfMatches' in r.text:
                m = re.search(r'<numOfMatches>(\d+)</numOfMatches>', r.text)
                if m: return int(m.group(1)) > 0
        r_track, _ = self.get(f'/ISAPI/ContentMgmt/record/tracks/{tid}')
        track_ok = (r_track and r_track.status_code == 200 and
                    f'<id>{tid}</id>' in r_track.text)
        if track_ok:
            r_hdd, _ = self.get('/ISAPI/ContentMgmt/Storage/hdd/1')
            if r_hdd and r_hdd.status_code == 200:
                low = r_hdd.text.lower()
                if '<status>ok</status>' in low: return True
                if '<status>error</status>' in low: return False
        if r_track and r_track.status_code == 200:
            low = r_track.text.lower()
            r2, _ = self.get('/ISAPI/ContentMgmt/InputProxy/channels')
            is_nvr = r2 and r2.status_code == 200 and 'InputProxyChannel' in r2.text
            if is_nvr: return True
            if '<enable>true</enable>'  in low: return True
            if '<enable>false</enable>' in low: return False
        return None

    def save_snap(self, img_bytes, ch, name):
        if not img_bytes: return
        d = os.path.join(SNAP_DIR, self.name.replace(' ', '_'))
        os.makedirs(d, exist_ok=True)
        safe = re.sub(r'[^\w\-_]', '_', name)
        with open(os.path.join(d, f'Ch{ch:02d}_{safe}.jpg'), 'wb') as f:
            f.write(img_bytes)

    # ── Full check ─────────────────────────────────────────────
    def check_all(self):
        res = {'name': self.name, 'ip': f'{self.ip}:{self.port}',
               'nvr_time': None, 'nvr_time_str': 'N/A',
               'time_ok': False, 'cameras': [], 'total': 0,
               'ok_count': 0, 'issue_count': 0, 'error': None}

        print(f"\n  {'─'*60}")
        print(f"  {self.name}  ({self.ip}:{self.port})")
        print(f"  {'─'*60}")

        dt, err = self.get_system_time()
        if dt is None:
            res['error'] = err
            print(f"  ERROR: {err}")
            return res

        res['nvr_time']     = dt
        res['nvr_time_str'] = dt.strftime('%Y-%m-%d %H:%M:%S')
        res['time_ok']      = (dt.date() == date.today())
        dvr_date_str        = dt.strftime('%d-%m-%Y')
        dvr_time_str        = dt.strftime('%H:%M:%S')

        print(f"  DVR Time : {res['nvr_time_str']}  "
              f"({'OK' if res['time_ok'] else '*** DATE MISMATCH ***'})")

        cam_names = self.get_camera_names()
        if not cam_names:
            res['error'] = 'No cameras found'
            print(f"  ERROR: {res['error']}")
            return res

        if self.skip_channels:
            cam_names = {ch: n for ch, n in cam_names.items()
                         if ch not in self.skip_channels}
            print(f"  Cameras  : {len(cam_names)} "
                  f"(skipping {len(self.skip_channels)} unused)")
        else:
            print(f"  Cameras  : {len(cam_names)}")

        print(f"  {'Ch':>4}  {'Name':<26}  {'Clarity':<22}  {'Rec':<5}  Status")
        print(f"  {'─'*60}")

        for ch in sorted(cam_names):
            name = cam_names[ch]
            print(f"  {ch:>4}  {name[:26]:<26}  ", end='', flush=True)

            img, _        = self.get_snapshot(ch)
            clarity, det  = check_clarity(img)
            recording     = self.check_recording(ch)
            self.save_snap(img, ch, name)

            issues = []
            # These clarity statuses are real problems
            if clarity in ('NO VIDEO', 'VERY DARK', 'OVEREXPOSED',
                           'LENS BLOCKED', 'BLURRY', 'NO SNAPSHOT',
                           'CHECK ERROR'):
                issues.append(clarity)
            # Night vision is informational — not an error
            # No timestamp is a warning
            if clarity == 'NO TIMESTAMP':
                issues.append('NO TIMESTAMP')
            if recording is False:
                issues.append('NO RECORDING')
            if not res['time_ok']:
                issues.append('DVR DATE WRONG')

            status = 'OK' if not issues else ' | '.join(issues)
            rec_s  = 'YES' if recording is True else 'NO' if recording is False else '?'
            print(f"{clarity:<22}  {rec_s:<5}  {status}")

            res['cameras'].append({
                'ch'       : ch,
                'name'     : name,
                'clarity'  : clarity,
                'detail'   : det,
                'cam_date' : dvr_date_str,
                'cam_time' : dvr_time_str,
                'date_ok'  : res['time_ok'],
                'recording': recording,
                'status'   : status,
            })

        res['total']       = len(res['cameras'])
        res['ok_count']    = sum(1 for c in res['cameras']
                                 if c['status'] in ('OK',) or
                                 c['clarity'] == 'NIGHT VISION')
        res['issue_count'] = res['total'] - res['ok_count']
        print(f"\n  Total:{res['total']}  OK:{res['ok_count']}  "
              f"Issues:{res['issue_count']}")
        return res


# ══════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════

def _fill(h): return PatternFill(start_color=h, end_color=h, fill_type='solid')
def _bdr():
    s = Side(style='thin', color='DDDDDD')
    return Border(left=s, right=s, top=s, bottom=s)
def _f(bold=False, size=9, color='000000'):
    return Font(name='Arial', bold=bold, size=size, color=color)


def build_excel(results):
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = 'Summary'
    _write_summary(ws_sum, results)
    ws_all = wb.create_sheet('All Cameras')
    _write_all_cameras(ws_all, results)
    wb.save(EXCEL_PATH)
    print(f"\n  Excel saved → {EXCEL_PATH}")


def _write_summary(ws, results):
    for col, w in zip('ABCDEFGHI', [18, 20, 22, 10, 8, 8, 8, 10, 18]):
        ws.column_dimensions[col].width = w

    c = ws.cell(row=1, column=1,
        value=f"CCTV Daily Report — {RPT_DATE}  |  {datetime.now().strftime('%H:%M:%S')}")
    c.font = _f(True, 13, 'FFFFFF'); c.fill = _fill(C['header'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:I1'); ws.row_dimensions[1].height = 26

    tc   = sum(r.get('total', 0)       for r in results)
    tok  = sum(r.get('ok_count', 0)    for r in results)
    tis  = sum(r.get('issue_count', 0) for r in results)
    tnv  = sum(sum(1 for c in r.get('cameras', [])
                   if c['clarity'] in ('NO VIDEO','VERY DARK','NO SNAPSHOT'))
               for r in results)
    tbl  = sum(sum(1 for c in r.get('cameras', [])
                   if c['clarity'] in ('BLURRY','LENS BLOCKED','OVEREXPOSED'))
               for r in results)
    tnr  = sum(sum(1 for c in r.get('cameras', [])
                   if c.get('recording') is False) for r in results)
    terr = sum(1 for r in results if r.get('error'))

    stats = [
        ('DVRs',        len(results), C['blue_hdr']),
        ('Cameras',     tc,           C['blue_hdr']),
        ('All OK',      tok,          C['ok_fg']),
        ('Issues',      tis,          C['err_fg'] if tis  else C['ok_fg']),
        ('No Video',    tnv,          C['err_fg'] if tnv  else C['ok_fg']),
        ('Blurry',      tbl,          C['err_fg'] if tbl  else C['ok_fg']),
        ('No Rec',      tnr,          C['err_fg'] if tnr  else C['ok_fg']),
        ('DVR Errors',  terr,         C['err_fg'] if terr else C['ok_fg']),
    ]
    for ci, (lbl, val, col) in enumerate(stats, 1):
        lc = ws.cell(row=2, column=ci, value=lbl)
        lc.font = _f(True, 8, 'FFFFFF')
        lc.fill = _fill(col)
        lc.alignment = Alignment(horizontal='center', vertical='center')
        vc = ws.cell(row=3, column=ci, value=val)
        vc.font = Font(name='Arial', bold=True, size=18, color=col)
        vc.fill = _fill(C['grey_bg'])
        vc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 32

    hdrs = ['DVR / NVR', 'IP Address', 'DVR Date & Time', 'Date OK?',
            'Total', 'OK', 'Issues', 'Night Vis', 'Status']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = _f(True, 10, 'FFFFFF')
        c.fill = _fill(C['blue_hdr'])
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = _bdr()
    ws.row_dimensions[4].height = 18
    ws.freeze_panes = 'A5'

    for ri, res in enumerate(results, 5):
        ws.row_dimensions[ri].height = 16
        err = res.get('error')
        bg  = (C['err_bg'] if err else
               C['ok_bg'] if not res.get('issue_count') else C['warn_bg'])
        night = sum(1 for c in res.get('cameras', [])
                    if 'NIGHT VISION' in c.get('clarity', ''))
        vals = (
            [res['name'], res['ip'], 'UNREACHABLE', '—', '—', '—', '—', '—', err]
            if err else
            [res['name'], res['ip'], res.get('nvr_time_str', '—'),
             'YES' if res.get('time_ok') else 'NO',
             res.get('total', 0), res.get('ok_count', 0),
             res.get('issue_count', 0), night,
             'ALL OK' if not res.get('issue_count')
             else f"{res['issue_count']} ISSUES"]
        )
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _f(size=9); c.fill = _fill(bg); c.border = _bdr()
            c.alignment = Alignment(horizontal='center', vertical='center')
            if ci == 1:
                c.alignment = Alignment(horizontal='left',
                                        vertical='center', indent=1)
            if ci == 9 and not err:
                fg = C['ok_fg'] if not res.get('issue_count') else C['err_fg']
                c.font = _f(True, 9, fg)


def _write_all_cameras(ws, results):
    col_widths = {'A': 16, 'B': 6, 'C': 30, 'D': 14,
                  'E': 12, 'F': 22, 'G': 12, 'H': 26}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    c = ws.cell(row=1, column=1,
        value=f"All Cameras — {RPT_DATE}  |  "
              f"Generated: {datetime.now().strftime('%H:%M:%S')}")
    c.font = _f(True, 12, 'FFFFFF'); c.fill = _fill(C['header'])
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells('A1:H1'); ws.row_dimensions[1].height = 24

    hdrs = ['DVR / NVR', 'Ch', 'Camera Name', 'DVR Date',
            'DVR Time', 'View Clarity', 'Recording', 'Status']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _f(True, 10, 'FFFFFF'); c.fill = _fill(C['blue_hdr'])
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = _bdr()
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = 'A2:H2'

    # Clarity colour map
    CLARITY_BG = {
        'CLEAR'                   : None,       # uses alternating green
        'NIGHT VISION'            : 'EEF2FF',   # light blue — informational
        'NIGHT VISION - NO TIMESTAMP': 'FFECC6', # orange — warn
        'NO TIMESTAMP'            : 'FFECC6',   # orange — warn
        'BLURRY'                  : 'FFDCE1',   # red
        'LENS BLOCKED'            : 'FFDCE1',   # red
        'OVEREXPOSED'             : 'FFDCE1',   # red
        'VERY DARK'               : 'FFDCE1',   # red
        'NO VIDEO'                : 'FFDCE1',   # red
        'NO SNAPSHOT'             : 'FFDCE1',   # red
        'CHECK ERROR'             : 'FFDCE1',   # red
    }

    row_i = 3
    for res in results:
        err = res.get('error')
        if err:
            ws.row_dimensions[row_i].height = 16
            vals = [res['name'], '—', f'ERROR: {err}',
                    '—', '—', 'UNREACHABLE', '—', err]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row_i, column=ci, value=val)
                c.font = _f(True, 9, C['err_fg'])
                c.fill = _fill(C['err_bg']); c.border = _bdr()
                c.alignment = Alignment(horizontal='center', vertical='center')
                if ci in (1, 3):
                    c.alignment = Alignment(horizontal='left',
                                            vertical='center', indent=1)
            row_i += 1
            continue

        for cam in res.get('cameras', []):
            ws.row_dimensions[row_i].height = 15
            clarity  = cam['clarity']
            rec      = cam['recording']
            status   = cam['status']
            is_alt   = (row_i % 2 == 0)

            # Background colour
            if clarity in CLARITY_BG and CLARITY_BG[clarity]:
                bg = CLARITY_BG[clarity]
            elif status != 'OK':
                bg = C['warn_bg']
            else:
                bg = C['ok_alt'] if is_alt else C['ok_bg']

            rec_s = ('YES' if rec is True else
                     'NO'  if rec is False else '?')
            vals = [
                res['name'],
                cam['ch'],
                cam['name'],
                cam['cam_date'],
                cam['cam_time'],
                clarity,
                rec_s,
                status,
            ]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row_i, column=ci, value=val)
                c.fill = _fill(bg); c.border = _bdr()
                c.font = _f(size=9)
                c.alignment = Alignment(horizontal='center', vertical='center')
                if ci in (1, 3):
                    c.alignment = Alignment(horizontal='left',
                                            vertical='center', indent=1)
                if ci == 6:   # Clarity
                    bad = clarity not in ('CLEAR', 'NIGHT VISION')
                    c.font = _f(True, 9,
                                C['err_fg'] if bad else
                                '0C447C'    if 'NIGHT' in clarity else
                                C['ok_fg'])
                if ci == 7:   # Recording
                    c.font = _f(True, 9,
                                C['ok_fg']   if rec is True else
                                C['err_fg']  if rec is False else
                                C['warn_fg'])
                if ci == 8:   # Status
                    c.font = _f(True, 9,
                                C['ok_fg'] if status == 'OK' else C['err_fg'])

            row_i += 1

    ws.auto_filter.ref = f'A2:H{row_i - 1}'


# ══════════════════════════════════════════════════════════════
def main():
    print('=' * 62)
    print(f'  CCTV Daily Check Bot v9 — {RPT_DATE}')
    print('=' * 62)
    with open(CONFIG_PATH) as f:
        config = json.load(f)
    devices = [d for d in config['dvr_nvr_list'] if d.get('enabled', True)]
    print(f'  {len(devices)} devices | Snapshots → {SNAP_DIR}')
    os.makedirs(SNAP_DIR, exist_ok=True)
    results = []
    for i, dev in enumerate(devices, 1):
        print(f'\n  [{i:02d}/{len(devices):02d}] {dev["name"]}', flush=True)
        results.append(DVRChecker(dev).check_all())
    print('\n\n  Building Excel report...')
    build_excel(results)
    tc  = sum(r.get('total', 0)       for r in results)
    ti  = sum(r.get('issue_count', 0) for r in results)
    te  = sum(1 for r in results if r.get('error'))
    print(f'\n{"=" * 62}')
    print(f'  DONE — Cameras:{tc}  Issues:{ti}  Errors:{te}')
    print(f'  {EXCEL_PATH}')
    print('=' * 62)

if __name__ == '__main__':
    main()
