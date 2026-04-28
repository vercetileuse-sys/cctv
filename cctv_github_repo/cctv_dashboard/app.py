"""CCTV Web Dashboard — Flask Backend with Login Authentication"""
from flask import Flask,render_template,jsonify,request,send_file,abort,session,redirect,url_for
import requests,json,os,re,threading,sqlite3,base64
import xml.etree.ElementTree as ET
from datetime import datetime,date
from requests.auth import HTTPDigestAuth,HTTPBasicAuth
import urllib3,io

try:
    import cv2,numpy as np
    CV2_OK=True
except: CV2_OK=False

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR,'nvr_config.json')
DB_PATH     = os.path.join(BASE_DIR,'db','cctv.db')
SNAP_BASE   = os.path.join(BASE_DIR,'snapshots')
TIMEOUT     = 12

app = Flask(__name__)
app.secret_key = 'cctv-aura-secret-2024-xK9mP'

# ── Login credentials ─────────────────────────────────────────
VALID_USER = 'Admin'
VALID_PASS = 'Auracctv#2024'

scan_state = {
    'running':False,'progress':0,'total':0,
    'current':'','results':[],'last_scan':None,'log':[]
}

# ── Auth helpers ──────────────────────────────────────────────
def logged_in():
    return session.get('auth') == True

def require_login(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args,**kwargs):
        if not logged_in():
            if request.path.startswith('/api/'):
                return jsonify({'error':'Not authenticated','login':True}),401
            return redirect(url_for('login_page'))
        return f(*args,**kwargs)
    return wrapper

# ── Database ──────────────────────────────────────────────────
def get_db():
    os.makedirs(os.path.dirname(DB_PATH),exist_ok=True)
    con = sqlite3.connect(DB_PATH,check_same_thread=False)
    con.row_factory = sqlite3.Row
    return con

def init_db():
    con=get_db()
    con.executescript("""
    CREATE TABLE IF NOT EXISTS scans(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scan_date TEXT NOT NULL,scan_time TEXT NOT NULL,
        total_dvrs INTEGER DEFAULT 0,total_cams INTEGER DEFAULT 0,
        ok_count INTEGER DEFAULT 0,issue_count INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS dvr_results(
        id INTEGER PRIMARY KEY AUTOINCREMENT,scan_id INTEGER,
        dvr_name TEXT,ip_port TEXT,nvr_time TEXT,time_ok INTEGER DEFAULT 0,
        total_cams INTEGER DEFAULT 0,ok_count INTEGER DEFAULT 0,
        issue_count INTEGER DEFAULT 0,error TEXT
    );
    CREATE TABLE IF NOT EXISTS camera_results(
        id INTEGER PRIMARY KEY AUTOINCREMENT,scan_id INTEGER,dvr_id INTEGER,
        dvr_name TEXT,channel INTEGER,cam_name TEXT,
        clarity TEXT,recording TEXT,cam_date TEXT,cam_time TEXT,
        status TEXT,snap_rel_path TEXT
    );
    """)
    con.commit(); con.close()

def save_to_db(results,scan_date,scan_time):
    try:
        con=get_db(); cur=con.cursor()
        tc=sum(len(r.get('cameras',[])) for r in results)
        tok=sum(r.get('ok',0) for r in results)
        tis=sum(r.get('issues',0) for r in results)
        cur.execute("INSERT INTO scans(scan_date,scan_time,total_dvrs,total_cams,ok_count,issue_count) VALUES(?,?,?,?,?,?)",
                    (scan_date,scan_time,len(results),tc,tok,tis))
        sid=cur.lastrowid
        for res in results:
            cur.execute("INSERT INTO dvr_results(scan_id,dvr_name,ip_port,nvr_time,time_ok,total_cams,ok_count,issue_count,error) VALUES(?,?,?,?,?,?,?,?,?)",
                        (sid,res['name'],res['ip'],res.get('nvr_time'),1 if res.get('time_ok') else 0,
                         res.get('total',0),res.get('ok',0),res.get('issues',0),res.get('error')))
            did=cur.lastrowid
            for cam in res.get('cameras',[]):
                cur.execute("INSERT INTO camera_results(scan_id,dvr_id,dvr_name,channel,cam_name,clarity,recording,cam_date,cam_time,status,snap_rel_path) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                            (sid,did,res['name'],cam['ch'],cam['name'],cam['clarity'],cam['recording'],
                             cam['date'],cam['time'],cam['status'],cam.get('snap_rel_path','')))
        con.commit(); con.close()
        print(f"[DB] Scan #{sid} saved — {tc} cameras")
        return sid
    except Exception as e:
        print(f"[DB ERROR] {e}")
        import traceback; traceback.print_exc()
        return None

# ── Camera helpers ────────────────────────────────────────────
def xfind(el,tag):
    v=el.find(tag)
    if v is not None: return v.text.strip() if v.text else ''
    for ns in ['{http://www.hikvision.com/ver20/XMLSchema}','{http://www.isapi.org/ver20/XMLSchema}']:
        v=el.find(f'{ns}{tag}')
        if v is not None: return v.text.strip() if v.text else ''
    for child in el.iter():
        if (child.tag.split('}')[-1] if '}' in child.tag else child.tag)==tag:
            return child.text.strip() if child.text else ''
    return None

def xfindall(el,tag):
    res=[]
    for child in el.iter():
        if (child.tag.split('}')[-1] if '}' in child.tag else child.tag)==tag: res.append(child)
    return res

def parse_hik_time(s):
    if not s: return None
    s=re.sub(r'[+-]\d{2}:\d{2}$','',s.strip()).replace('Z','').strip()[:19]
    try: return datetime.strptime(s,'%Y-%m-%dT%H:%M:%S')
    except: return None

def check_clarity(img_bytes):
    if not img_bytes or len(img_bytes)<3000: return 'NO SNAPSHOT',''
    if not CV2_OK: return ('CLEAR' if len(img_bytes)>10000 else 'POSSIBLY DARK','')
    try:
        arr=cv2.imdecode(np.frombuffer(img_bytes,np.uint8),cv2.IMREAD_COLOR)
        if arr is None: return 'NO SNAPSHOT',''
        h,w=arr.shape[:2]; gray=cv2.cvtColor(arr,cv2.COLOR_BGR2GRAY)
        mean_bri=float(np.mean(gray))
        if mean_bri<12: return 'NO VIDEO',f'{mean_bri:.0f}'
        if mean_bri<38: return 'VERY DARK',f'{mean_bri:.0f}'
        if mean_bri>240: return 'OVEREXPOSED',f'{mean_bri:.0f}'
        b_m=float(np.mean(arr[:,:,0])); g_m=float(np.mean(arr[:,:,1])); r_m=float(np.mean(arr[:,:,2]))
        is_night=max(abs(b_m-g_m),abs(g_m-r_m),abs(b_m-r_m))<6.0
        if float(np.std(gray))<8.0: return 'LENS BLOCKED',''
        cy1,cy2,cx1,cx2=h//4,3*h//4,w//4,3*w//4
        center=gray[cy1:cy2,cx1:cx2]
        lap_var=float(cv2.Laplacian(center,cv2.CV_64F).var())
        edges=cv2.Canny(center,50,150)
        edge_pct=float(np.count_nonzero(edges))/center.size*100
        ts_region=gray[:int(h*0.10),:int(w*0.40)]
        ts_vis=np.sum(ts_region>200)/ts_region.size*100>0.3
        if lap_var<60 and edge_pct<1.0: return 'BLURRY',f'{lap_var:.0f}'
        if lap_var<35: return 'BLURRY',f'{lap_var:.0f}'
        if is_night: return 'NIGHT VISION',f'{mean_bri:.0f}'
        if not ts_vis: return 'NO TIMESTAMP',f'{mean_bri:.0f}'
        return 'CLEAR',f'{mean_bri:.0f}'
    except Exception as e: return 'ERROR',str(e)

class DVRChecker:
    def __init__(self,cfg):
        self.name=cfg['name']; self.ip=cfg['ip_address']
        self.port=cfg.get('port',80); self.user=cfg['username']; self.pwd=cfg['password']
        self.base=f"http://{self.ip}:{self.port}"; self.auth=HTTPDigestAuth(self.user,self.pwd)
        self.skip=set(cfg.get('skip_channels',[]))
    def get(self,path,stream=False):
        url=f"{self.base}{path}"
        for auth in [HTTPDigestAuth(self.user,self.pwd),HTTPBasicAuth(self.user,self.pwd)]:
            try:
                r=requests.get(url,auth=auth,verify=False,timeout=TIMEOUT,stream=stream)
                if r.status_code==401: continue
                self.auth=auth; return r,None
            except requests.exceptions.ConnectTimeout: return None,'TIMEOUT'
            except requests.exceptions.ConnectionError: return None,'UNREACHABLE'
            except Exception as e: return None,str(e)
        return None,'AUTH_FAILED'
    def post(self,path,body):
        try:
            r=requests.post(f"{self.base}{path}",data=body,auth=self.auth,
                            headers={'Content-Type':'application/xml'},verify=False,timeout=TIMEOUT)
            return r,None
        except Exception as e: return None,str(e)
    def system_time(self):
        r,err=self.get('/ISAPI/System/time')
        if err: return None,err
        if r.status_code!=200: return None,f'HTTP {r.status_code}'
        try: root=ET.fromstring(r.text)
        except: return None,'XML error'
        t=xfind(root,'localTime')
        if not t: return None,'localTime not found'
        dt=parse_hik_time(t)
        return (dt,'OK') if dt else (None,f'Parse error: {t}')
    def cam_names(self):
        names={}
        r,_=self.get('/ISAPI/System/Video/inputs/channels')
        if r and r.status_code==200:
            try:
                root=ET.fromstring(r.text)
                for ch in xfindall(root,'VideoInputChannel'):
                    cid=xfind(ch,'id'); nm=xfind(ch,'n') or xfind(ch,'name')
                    if cid and nm: names[int(cid)]=nm
            except: pass
        if not names:
            r,_=self.get('/ISAPI/ContentMgmt/InputProxy/channels')
            if r and r.status_code==200:
                try:
                    root=ET.fromstring(r.text)
                    for ch in xfindall(root,'InputProxyChannel'):
                        cid=xfind(ch,'id'); nm=xfind(ch,'name')
                        if cid and nm: names[int(cid)]=nm
                except: pass
        return names
    def snapshot(self,ch):
        tid=ch*100+1
        for path in [f'/ISAPI/Streaming/channels/{tid}/picture',f'/ISAPI/Streaming/channels/{ch}/picture']:
            r,err=self.get(path,stream=True)
            if err: continue
            if r and r.status_code==200 and len(r.content)>2000: return r.content,None
        return None,'FAILED'
    def recording(self,ch):
        tid=ch*100+1; today=date.today().strftime('%Y-%m-%d')
        for tz in ['+05:30','Z']:
            body=(f'<CMSearchDescription><searchID>{ch}</searchID>'
                  f'<trackList><trackID>{tid}</trackID></trackList>'
                  f'<timeSpanList><timeSpan>'
                  f'<startTime>{today}T00:00:01{tz}</startTime>'
                  f'<endTime>{today}T23:59:59{tz}</endTime>'
                  f'</timeSpan></timeSpanList>'
                  f'<maxResults>1</maxResults><searchResultPostion>0</searchResultPostion>'
                  f'<metadataList><metadataDescriptor>'
                  f'//recordType.meta.hikvision.com/dataType'
                  f'</metadataDescriptor></metadataList></CMSearchDescription>')
            r,_=self.post('/ISAPI/ContentMgmt/search',body)
            if r and r.status_code==200 and 'numOfMatches' in r.text:
                m=re.search(r'<numOfMatches>(\d+)</numOfMatches>',r.text)
                if m: return int(m.group(1))>0
        rt,_=self.get(f'/ISAPI/ContentMgmt/record/tracks/{tid}')
        if rt and rt.status_code==200:
            if f'<id>{tid}</id>' in rt.text:
                rh,_=self.get('/ISAPI/ContentMgmt/Storage/hdd/1')
                if rh and '<status>ok</status>' in rh.text.lower(): return True
            low=rt.text.lower()
            r2,_=self.get('/ISAPI/ContentMgmt/InputProxy/channels')
            if r2 and 'InputProxyChannel' in r2.text: return True
            if '<enable>true</enable>' in low: return True
            if '<enable>false</enable>' in low: return False
        return None
    def check_all(self,snap_dir):
        res={'name':self.name,'ip':f'{self.ip}:{self.port}',
             'raw_ip':self.ip,'raw_port':self.port,
             'user':self.user,'pwd':self.pwd,
             'nvr_time':None,'time_ok':False,
             'cameras':[],'total':0,'ok':0,'issues':0,'error':None}
        dt,err=self.system_time()
        if dt is None: res['error']=err; return res
        res['nvr_time']=dt.strftime('%d-%m-%Y %H:%M:%S')
        res['time_ok']=(dt.date()==date.today())
        ddate=dt.strftime('%d-%m-%Y'); dtime=dt.strftime('%H:%M:%S')
        names=self.cam_names()
        if not names: res['error']='No cameras found'; return res
        if self.skip: names={ch:n for ch,n in names.items() if ch not in self.skip}
        dvr_safe=re.sub(r'[^\w\-_]','_',self.name)
        dvr_dir=os.path.join(snap_dir,dvr_safe)
        os.makedirs(dvr_dir,exist_ok=True)
        for ch in sorted(names):
            name=names[ch]
            img,_=self.snapshot(ch)
            clarity,_=check_clarity(img)
            rec=self.recording(ch)
            snap_rel=''
            if img:
                safe=re.sub(r'[^\w\-_]','_',name)
                fname=f'Ch{ch:02d}_{safe}.jpg'
                abs_path=os.path.join(dvr_dir,fname)
                with open(abs_path,'wb') as f: f.write(img)
                date_folder=os.path.basename(snap_dir)
                snap_rel=f"{date_folder}/{dvr_safe}/{fname}"
            issues=[]
            if clarity not in ('CLEAR','NIGHT VISION'): issues.append(clarity)
            if rec is False: issues.append('NO RECORDING')
            if not res['time_ok']: issues.append('DVR DATE WRONG')
            status='OK' if not issues else ' | '.join(issues)
            rec_s='YES' if rec is True else 'NO' if rec is False else '?'
            res['cameras'].append({'ch':ch,'name':name,'clarity':clarity,'recording':rec_s,
                                   'date':ddate,'time':dtime,'status':status,'snap_rel_path':snap_rel})
        res['total']=len(res['cameras'])
        res['ok']=sum(1 for c in res['cameras'] if c['status']=='OK' or c['clarity']=='NIGHT VISION')
        res['issues']=res['total']-res['ok']
        return res

def run_scan():
    global scan_state
    with open(CONFIG_PATH) as f: config=json.load(f)
    devices=[d for d in config['dvr_nvr_list'] if d.get('enabled',True)]
    now=datetime.now()
    scan_date=now.strftime('%Y-%m-%d'); scan_time=now.strftime('%H:%M:%S')
    scan_dir=os.path.join(SNAP_BASE,scan_date)
    os.makedirs(scan_dir,exist_ok=True)
    scan_state.update({'running':True,'progress':0,'total':len(devices),'results':[],'log':[],'current':''})
    for i,dev in enumerate(devices):
        scan_state['current']=dev['name']
        scan_state['log'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Checking {dev['name']}...")
        result=DVRChecker(dev).check_all(scan_dir)
        scan_state['results'].append(result)
        scan_state['progress']=i+1
        msg=f"✗ {result['error']}" if result.get('error') else f"✓ {result['ok']}/{result['total']} OK, {result['issues']} issues"
        scan_state['log'].append(f"[{datetime.now().strftime('%H:%M:%S')}] {dev['name']}: {msg}")
    sid=save_to_db(scan_state['results'],scan_date,scan_time)
    if sid:
        scan_state['log'].append(f"[{datetime.now().strftime('%H:%M:%S')}] ✓ Saved to database — Scan #{sid}")
    else:
        scan_state['log'].append(f"[{datetime.now().strftime('%H:%M:%S')}] ✗ Database save failed")
    scan_state.update({'running':False,'last_scan':now.strftime('%d-%m-%Y %H:%M:%S'),'current':''})

# ══════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════

@app.route('/login', methods=['GET','POST'])
def login_page():
    error=''
    if request.method=='POST':
        u=request.form.get('username','')
        p=request.form.get('password','')
        if u==VALID_USER and p==VALID_PASS:
            session['auth']=True
            session.permanent=True
            return redirect(url_for('index'))
        error='Invalid username or password'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/')
@require_login
def index(): return render_template('index.html')

@app.route('/api/start_scan',methods=['POST'])
@require_login
def start_scan():
    if scan_state['running']: return jsonify({'error':'Already running'}),400
    threading.Thread(target=run_scan,daemon=True).start()
    return jsonify({'status':'started'})

@app.route('/api/scan_status')
@require_login
def scan_status():
    return jsonify({'running':scan_state['running'],'progress':scan_state['progress'],
                    'total':scan_state['total'],'current':scan_state['current'],
                    'last_scan':scan_state['last_scan'],'log':scan_state['log'][-40:]})

@app.route('/api/results')
@require_login
def get_results(): return jsonify(scan_state['results'])

@app.route('/api/snapshot')
@require_login
def get_snapshot():
    rel=request.args.get('rel','')
    if not rel: abort(404)
    rel=rel.lstrip('/').replace('..','')
    full=os.path.join(SNAP_BASE,rel)
    if not os.path.exists(full): abort(404)
    return send_file(full,mimetype='image/jpeg',max_age=3600)

@app.route('/api/history')
@require_login
def get_history():
    con=get_db()
    rows=con.execute("SELECT * FROM scans ORDER BY id DESC LIMIT 90").fetchall()
    con.close(); return jsonify([dict(r) for r in rows])

@app.route('/api/history/<int:sid>')
@require_login
def get_history_detail(sid):
    con=get_db()
    scan=con.execute("SELECT * FROM scans WHERE id=?",(sid,)).fetchone()
    dvrs=con.execute("SELECT * FROM dvr_results WHERE scan_id=? ORDER BY id",(sid,)).fetchall()
    cams=con.execute("SELECT * FROM camera_results WHERE scan_id=? ORDER BY dvr_name,channel",(sid,)).fetchall()
    con.close()
    return jsonify({'scan':dict(scan) if scan else {},'dvrs':[dict(r) for r in dvrs],'cameras':[dict(r) for r in cams]})

@app.route('/api/history/<int:sid>/excel')
@require_login
def history_excel(sid):
    import openpyxl
    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    con=get_db()
    scan=con.execute("SELECT * FROM scans WHERE id=?",(sid,)).fetchone()
    if not scan: con.close(); return jsonify({'error':'Not found'}),404
    dvrs=con.execute("SELECT * FROM dvr_results WHERE scan_id=?",(sid,)).fetchall()
    cams=con.execute("SELECT * FROM camera_results WHERE scan_id=? ORDER BY dvr_name,channel",(sid,)).fetchall()
    con.close()
    scan=dict(scan); dvrs=[dict(d) for d in dvrs]; cams=[dict(c) for c in cams]
    wb=openpyxl.Workbook()
    def fill(h): return PatternFill(start_color=h,end_color=h,fill_type='solid')
    def bdr():
        s=Side(style='thin',color='DDDDDD'); return Border(left=s,right=s,top=s,bottom=s)
    ws=wb.active; ws.title='Summary'
    c=ws.cell(row=1,column=1,value=f"CCTV Report — {scan['scan_date']} {scan['scan_time']}")
    c.font=Font(name='Arial',bold=True,size=13,color='FFFFFF')
    c.fill=fill('1a1f3a'); ws.merge_cells('A1:H1')
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=26
    for ci,(h,w2) in enumerate(zip(['DVR/NVR','IP','DVR Time','Date OK','Total','OK','Issues','Error'],
                                    [16,20,22,10,8,8,8,30]),1):
        ws.column_dimensions[get_column_letter(ci)].width=w2
        c=ws.cell(row=2,column=ci,value=h)
        c.font=Font(name='Arial',bold=True,size=10,color='FFFFFF')
        c.fill=fill('2e4057'); c.border=bdr()
        c.alignment=Alignment(horizontal='center',vertical='center')
    for ri,d in enumerate(dvrs,3):
        bg='FFDCE1' if d['error'] else ('E8F5E9' if ri%2==0 else 'F1F8E9')
        vals=[d['dvr_name'],d['ip_port'],d['nvr_time'] or '','YES' if d['time_ok'] else 'NO',
              d['total_cams'],d['ok_count'],d['issue_count'],d['error'] or '']
        for ci,val in enumerate(vals,1):
            c=ws.cell(row=ri,column=ci,value=val)
            c.fill=fill(bg); c.border=bdr(); c.font=Font(name='Arial',size=9)
            c.alignment=Alignment(horizontal='center',vertical='center')
        ws.row_dimensions[ri].height=15
    ws2=wb.create_sheet('All Cameras')
    for ci,(h,w2) in enumerate(zip(['DVR/NVR','Ch','Camera Name','Date','Time','Clarity','Recording','Status'],
                                    [16,6,30,14,12,22,12,28]),1):
        ws2.column_dimensions[get_column_letter(ci)].width=w2
        c=ws2.cell(row=1,column=ci,value=h)
        c.font=Font(name='Arial',bold=True,size=10,color='FFFFFF')
        c.fill=fill('1a1f3a'); c.border=bdr()
        c.alignment=Alignment(horizontal='center',vertical='center')
    ws2.freeze_panes='A2'
    for ri,cam in enumerate(cams,2):
        bg=('FFDCE1' if cam['status']!='OK' and 'NIGHT' not in str(cam['clarity'])
            else ('E8F5E9' if ri%2==0 else 'F1F8E9'))
        vals=[cam['dvr_name'],cam['channel'],cam['cam_name'],cam['cam_date'],
              cam['cam_time'],cam['clarity'],cam['recording'],cam['status']]
        for ci,val in enumerate(vals,1):
            c=ws2.cell(row=ri,column=ci,value=val)
            c.fill=fill(bg); c.border=bdr(); c.font=Font(name='Arial',size=9)
            c.alignment=Alignment(horizontal='left' if ci in(1,3) else 'center',
                                  vertical='center',indent=1 if ci in(1,3) else 0)
        ws2.row_dimensions[ri].height=15
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"CCTV_Report_{scan['scan_date'].replace('-','')}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/download_excel')
@require_login
def download_excel():
    import openpyxl
    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    results=scan_state['results']
    if not results: return jsonify({'error':'No results. Run a scan first.'}),400
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='CCTV Report'
    def fill(h): return PatternFill(start_color=h,end_color=h,fill_type='solid')
    def bdr():
        s=Side(style='thin',color='DDDDDD'); return Border(left=s,right=s,top=s,bottom=s)
    hdrs=['DVR/NVR','Ch','Camera Name','Date','Time','Clarity','Recording','Status']
    for ci,(h,w2) in enumerate(zip(hdrs,[16,6,30,14,12,22,12,28]),1):
        ws.column_dimensions[get_column_letter(ci)].width=w2
        c=ws.cell(row=1,column=ci,value=h)
        c.font=Font(name='Arial',bold=True,size=10,color='FFFFFF')
        c.fill=fill('1a1f3a'); c.border=bdr()
        c.alignment=Alignment(horizontal='center',vertical='center')
    ws.freeze_panes='A2'; row=2
    for res in results:
        if res.get('error'):
            for ci in range(1,9):
                c=ws.cell(row=row,column=ci,value=res['name'] if ci==1 else ('UNREACHABLE' if ci==6 else ''))
                c.font=Font(name='Arial',bold=True,color='9C0006')
                c.fill=fill('FFDCE1'); c.border=bdr()
            row+=1; continue
        for cam in res.get('cameras',[]):
            bg=('FFDCE1' if cam['status']!='OK' and 'NIGHT' not in cam['clarity']
                else ('E8F5E9' if row%2==0 else 'F1F8E9'))
            vals=[res['name'],cam['ch'],cam['name'],cam['date'],cam['time'],cam['clarity'],cam['recording'],cam['status']]
            for ci,val in enumerate(vals,1):
                c=ws.cell(row=row,column=ci,value=val)
                c.fill=fill(bg); c.border=bdr(); c.font=Font(name='Arial',size=9)
                c.alignment=Alignment(horizontal='left' if ci in(1,3) else 'center',
                                      vertical='center',indent=1 if ci in(1,3) else 0)
            ws.row_dimensions[row].height=15; row+=1
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,
                     download_name=f"CCTV_{datetime.now().strftime('%d-%m-%Y_%H%M')}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/config')
@require_login
def get_config():
    with open(CONFIG_PATH) as f: c=json.load(f)
    return jsonify(c['dvr_nvr_list'])

if __name__=='__main__':
    init_db()
    print("="*55)
    print("  CCTV Dashboard  →  http://localhost:5000")
    print(f"  Login: {VALID_USER} / {VALID_PASS}")
    print("="*55)
    app.run(debug=False,host='0.0.0.0',port=5000,threaded=True)
